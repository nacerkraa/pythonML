import openpyxl as xl

wb = xl.load_workbook("file.xlsx")

sheet = wb["Sheet1"]

cell = sheet["a1"]

cell = sheet.cell(1, 1)

# Create a for loop
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# save the information on new file
wb.save('file2.xlsx')